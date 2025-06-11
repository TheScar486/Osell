from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def distribution_home(request):
    user = request.user
    sucursal_nombre = request.session.get('sucursal_nombre', 'Sucursal no definida')

    context = {
        'user': user,
        'sucursal_nombre': sucursal_nombre,  
    }

    return render(request, 'distribution/distribution.html', context)

@login_required
def history_view(request):
    sucursal_nombre = request.session.get('sucursal_nombre', 'Sucursal no definida')

    context = {
        'title': 'Gestión de Historial',
        'user': request.user,
        'sucursal_nombre': sucursal_nombre, 
    }
    return render(request, 'distribution/sections/history.html', context)

@login_required
def inbound_transfers_view(request):
    sucursal_nombre = request.session.get('sucursal_nombre', 'Sucursal no definida')

    context = {
        'title': 'Gestión de Transferencias de Entrantes',
        'user': request.user,
        'sucursal_nombre': sucursal_nombre,
    }
    return render(request, 'distribution/sections/inbound_transfers.html', context)

@login_required
def transfers_view(request):
    sucursal_nombre = request.session.get('sucursal_nombre', 'Sucursal no definida')
    context = {
        'title': 'Gestión de Transferencias',
        'user': request.user,
        'sucursal_nombre': sucursal_nombre,
    }
    return render(request, 'distribution/sections/transfers.html', context)

@login_required
def requests_view(request):
    sucursal_nombre = request.session.get('sucursal_nombre', 'Sucursal no definida')
    context = {
        'title': 'Gestión de Solicitudes',
        'user': request.user,
        'sucursal_nombre': sucursal_nombre,
    }
    return render(request, 'distribution/sections/requests.html', context)

# views.py
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Catalogo
from .serializers import CatalogoSerializer

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def catalogo_activo(request):
    # Esto es para verificar si el usuario tiene sucursal_id
    print("User:", request.user)
    print("Sucursal ID:", getattr(request.user, 'sucursal_id', 'NO TIENE'))

    catalogos = Catalogo.objects.filter(Estado='ACTIVO')
    serializer = CatalogoSerializer(catalogos, many=True, context={'request': request})
    return Response(serializer.data)

from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import CatalogoUnidades, Catalogo
from .serializers import CatalogoUnidadesSerializer

@api_view(['GET'])
def obtener_factores(request):
    """
    Obtiene factores de conversión por código de artículo o ID
    """
    codigo_principal = request.query_params.get('codigo', None)
    id_articulo = request.query_params.get('id_articulo', None)
    
    if not codigo_principal and not id_articulo:
        return Response({"error": "Se requiere parámetro 'codigo' o 'id_articulo'"}, status=400)
    
    try:
        # Si nos dan código, primero buscamos el artículo
        if codigo_principal:
            articulo = Catalogo.objects.get(Codigo_Principal=codigo_principal)
            id_articulo = articulo.ID_Articulo
        
        factores = CatalogoUnidades.objects.filter(id_articulo=id_articulo)
        
        if not factores.exists():
            return Response({"message": "No se encontraron factores para este artículo"}, status=404)
            
        serializer = CatalogoUnidadesSerializer(factores, many=True)
        return Response(serializer.data)
        
    except Catalogo.DoesNotExist:
        return Response({"error": "Artículo no encontrado"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import SolicitudPedido
from apps.account.models import Empresa, Sucursal
from django.core.exceptions import ValidationError
from django.db.models import Max
from django.contrib.auth.decorators import login_required
import logging

logger = logging.getLogger(__name__)

def generar_folio_pedido():
    ultimo_folio = SolicitudPedido.objects.aggregate(max_folio=Max('pedido_id'))['max_folio']
    if ultimo_folio:
        try:
            numero = int(ultimo_folio.replace('PE-', ''))
            nuevo_numero = numero + 1
        except ValueError:
            nuevo_numero = 1
    else:
        nuevo_numero = 1
    return f"PE-{nuevo_numero:06d}"

@login_required
def crear_solicitud(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            datos_tabla = data.get('datos', [])

            if not datos_tabla:
                return JsonResponse({'success': False, 'error': 'No se recibieron datos'}, status=400)

            sucursal_id = request.session.get('sucursal_id')
            empresa_id = request.session.get('empresa_id')

            if not sucursal_id or not empresa_id:
                return JsonResponse({'success': False, 'error': 'Falta información de empresa/sucursal'}, status=400)

            sucursal = Sucursal.objects.get(id=sucursal_id)
            empresa = Empresa.objects.get(id=empresa_id)

            pedido_id = generar_folio_pedido()
            created_items = []

            for item in datos_tabla:
                solicitud = SolicitudPedido(
                    cantidad=item.get('cantidad', 0),
                    factor=item.get('factor'),
                    codigo=item.get('codigo', ''),
                    descripcion=item.get('descripcion', ''),
                    departamento=item.get('departamento', ''),
                    compra=item.get('compra'),
                    venta=item.get('venta'),
                    importe=item.get('importe'),
                    comentario=item.get('comentario', ''),
                    empresa=empresa,
                    sucursal=sucursal,
                    usuario=request.user,
                    pedido_id=pedido_id,
                    estado=item.get('estado', 'Pendiente')  # Establecer estado si se pasa en el item
                )
                solicitud.full_clean()  # Valida el modelo antes de guardar
                solicitud.save()
                created_items.append(solicitud.id)

            return JsonResponse({
                'success': True,
                'count': len(created_items),
                'ids': created_items,
                'pedido_id': pedido_id
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
        except Sucursal.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Sucursal no encontrada'}, status=404)
        except Empresa.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Empresa no encontrada'}, status=404)
        except ValidationError as e:
            return JsonResponse({'success': False, 'error': f'Error de validación: {dict(e)}'}, status=400)
        except Exception as e:
            logger.error(f"Error al crear solicitud: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import SolicitudPedido
from .serializers import SolicitudPedidoSerializer

class SolicitudPedidoListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Filtrar las solicitudes por empresa del usuario autenticado
        empresa_usuario = request.user.empresa  # Obtener la empresa del usuario
        pedidos = SolicitudPedido.objects.filter(usuario__empresa=empresa_usuario).order_by('-fecha_registro')

        # Agrupar manualmente por pedido_id para no mostrar las mismas solicitudes múltiples veces
        vistos = {}
        pedidos_unicos = []
        for p in pedidos:
            if p.pedido_id not in vistos:
                vistos[p.pedido_id] = True
                pedidos_unicos.append(p)

        serializer = SolicitudPedidoSerializer(pedidos_unicos, many=True)
        return Response(serializer.data)

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import SolicitudPedido
from apps.account.models import Usuario
from django.utils import timezone
import math

def export_solicitudes_excel(request, pedido_id=None):
    user = request.user

    if not pedido_id:
        return HttpResponse("Se requiere un ID de pedido para generar este formato de reporte.", status=400)

    solicitudes = SolicitudPedido.objects.filter(pedido_id=pedido_id, usuario__empresa=user.empresa).order_by('id')

    if not solicitudes.exists():
        return HttpResponse(f"No se encontraron solicitudes para el Pedido ID: {pedido_id} o no tiene permiso para acceder.", status=404)

    primer_solicitud = solicitudes.first()

    workbook = openpyxl.Workbook()
    
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # --- Configuración de Paginación ---
    HEADER_ROWS = 6
    FOOTER_ROWS = 2
    MAX_DATA_ROWS_PER_PAGE = 42

    total_items = solicitudes.count()
    total_pages = math.ceil(total_items / MAX_DATA_ROWS_PER_PAGE)

    fecha_generacion = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M:%S")

    # --- Definición de Márgenes en Centímetros y Conversión a Pulgadas ---
    MARGIN_TOP_CM = 1.91
    MARGIN_BOTTOM_CM = 1.1
    MARGIN_LEFT_CM = 0.64
    MARGIN_RIGHT_CM = 0.64
    MARGIN_HEADER_CM = 0.76
    MARGIN_FOOTER_CM = 0.76

    CM_TO_INCHES = 1 / 2.54

    margin_top_inches = MARGIN_TOP_CM * CM_TO_INCHES
    margin_bottom_inches = MARGIN_BOTTOM_CM * CM_TO_INCHES
    margin_left_inches = MARGIN_LEFT_CM * CM_TO_INCHES
    margin_right_inches = MARGIN_RIGHT_CM * CM_TO_INCHES
    margin_header_inches = MARGIN_HEADER_CM * CM_TO_INCHES
    margin_footer_inches = MARGIN_FOOTER_CM * CM_TO_INCHES


    # --- Definición de Anchos de Columna ---
    column_widths = {
        'A': 8.5,
        'B': 8.5,
        'C': 15,
        'D': 35,
        'E': 20,
        'F': 18
    }

    # --- Estilos adicionales ---
    # Encabezado de la tabla: Azul oscuro mate (fondo), Blanco (texto)
    header_fill = PatternFill(start_color="1F4068", end_color="1F4068", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Contenido de la tabla (filas de datos): Texto negro, fondo blanco
    data_row_font = Font(size=9, color="000000") # Texto NEGRO para visibilidad en fondo blanco

    # Color gris claro para las líneas discontinuas
    LIGHT_GRAY_COLOR = "D3D3D3" # Hex para LightGray

    # Bordes discontinuos: Top para todas las filas de datos, y Bottom para la última fila de datos
    dashed_top_border_style = Side(border_style='dashed', color=LIGHT_GRAY_COLOR)
    dashed_bottom_border_style = Side(border_style='dashed', color=LIGHT_GRAY_COLOR)


    # Iterar para crear cada página
    for page_num in range(1, total_pages + 1):
        sheet_name = f"Detalle Solicitud Pag {page_num}"
        sheet = workbook.create_sheet(title=sheet_name)

        # --- CABECERA (Se repite en cada página) ---

        # Título principal "SOLICITUD"
        sheet['A1'] = "SOLICITUD"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:F1')

        # Información de la solicitud
        row_offset_header = 2
        
        sheet[f'A{row_offset_header}'] = "Empresa"
        sheet[f'A{row_offset_header}'].font = Font(bold=True)
        sheet[f'B{row_offset_header}'] = primer_solicitud.empresa.nombre_legal if primer_solicitud.empresa else ""
        sheet[f'E{row_offset_header}'] = "Pedido ID"
        sheet[f'E{row_offset_header}'].font = Font(bold=True)
        sheet[f'F{row_offset_header}'] = primer_solicitud.pedido_id

        sheet[f'A{row_offset_header + 1}'] = "Sucursal"
        sheet[f'A{row_offset_header + 1}'].font = Font(bold=True)
        sheet[f'B{row_offset_header + 1}'] = primer_solicitud.sucursal.nombre if primer_solicitud.sucursal else ""
        sheet[f'E{row_offset_header + 1}'] = "Estado"
        sheet[f'E{row_offset_header + 1}'].font = Font(bold=True)
        sheet[f'F{row_offset_header + 1}'] = primer_solicitud.estado

        sheet[f'A{row_offset_header + 2}'] = "Usuario"
        sheet[f'A{row_offset_header + 2}'].font = Font(bold=True)
        sheet[f'B{row_offset_header + 2}'] = primer_solicitud.usuario.username if primer_solicitud.usuario else ""
        sheet[f'E{row_offset_header + 2}'] = "Fecha Registro"
        sheet[f'E{row_offset_header + 2}'].font = Font(bold=True)
        sheet[f'F{row_offset_header + 2}'] = primer_solicitud.fecha_registro.strftime("%Y-%m-%d %H:%M:%S") if primer_solicitud.fecha_registro else ""

        # --- Encabezados de la Tabla de Artículos (fila 6) ---
        header_table_row = 6
        
        header_cells_to_style = ['A', 'B', 'C', 'D', 'E', 'F']
        for col_letter in header_cells_to_style:
            cell = sheet[f'{col_letter}{header_table_row}']
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left')
            cell.fill = header_fill

        sheet[f'A{header_table_row}'] = "Cantidad"
        sheet[f'B{header_table_row}'] = "Factor"
        sheet[f'C{header_table_row}'] = "Código"
        sheet[f'D{header_table_row}'] = "Descripción"
        sheet.merge_cells(f'D{header_table_row}:E{header_table_row}')
        sheet[f'F{header_table_row}'] = "Departamento"

        # --- DATOS DE LA TABLA DE ARTÍCULOS ---
        start_index = (page_num - 1) * MAX_DATA_ROWS_PER_PAGE
        end_index = min(start_index + MAX_DATA_ROWS_PER_PAGE, total_items)
        
        current_data_row = header_table_row + 1

        for i in range(start_index, end_index):
            solicitud = solicitudes[i]
            
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = sheet[f"{col_letter}{current_data_row}"]
                cell.font = data_row_font 
                # No se aplica fill para dejar el fondo BLANCO
                cell.border = Border(top=dashed_top_border_style) # Aplicar solo borde superior

            sheet[f"A{current_data_row}"] = i + 1
            sheet[f"B{current_data_row}"] = solicitud.factor
            sheet[f"C{current_data_row}"] = solicitud.codigo
            sheet[f"D{current_data_row}"] = solicitud.descripcion
            sheet.merge_cells(f'D{current_data_row}:E{current_data_row}')
            sheet[f"F{current_data_row}"] = solicitud.departamento
            current_data_row += 1

        # --- Aplicar borde inferior a la última fila de datos de la página ---
        # last_data_row_on_page es la fila donde está el último item de la tabla en esta página
        last_data_row_on_page = header_table_row + (end_index - start_index)
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = sheet[f"{col_letter}{last_data_row_on_page}"]
            # Obtener el borde existente (que ya debería tener el top border)
            current_border = cell.border if cell.border else Border()
            # Crear un nuevo borde combinando el superior existente con el inferior
            new_border = Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=dashed_bottom_border_style # Añadir el borde inferior
            )
            cell.border = new_border


        # --- PIE DE PÁGINA (Se repite en cada página) ---
        footer_start_row = last_data_row_on_page + 2 

        # "Items XXX" (sin borde discontinuo)
        sheet[f'A{footer_start_row}'] = f"Items {total_items}"
        sheet[f'A{footer_start_row}'].font = Font(bold=True, color="000000") # Texto negro para Items

        # Fecha y hora de generación (sin borde discontinuo)
        sheet[f'C{footer_start_row}'] = f"fecha y hora de generacion {fecha_generacion}"
        sheet[f'C{footer_start_row}'].alignment = Alignment(horizontal='center')
        sheet.merge_cells(f'C{footer_start_row}:E{footer_start_row}')
        sheet[f'C{footer_start_row}'].font = Font(color="000000")

        # Página X de Y (sin borde discontinuo)
        sheet[f'F{footer_start_row}'] = f"PAG {page_num} de {total_pages}"
        sheet[f'F{footer_start_row}'].alignment = Alignment(horizontal='right')
        sheet[f'F{footer_start_row}'].font = Font(color="000000")

        # --- APLICAR ANCHO DE COLUMNAS PARA ESTA HOJA ---
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        # --- Configurar impresión: Ajuste de página y MÁRGENES ---
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToHeight = False
        sheet.page_setup.fitToWidth = 1
        
        sheet.page_margins.left = margin_left_inches
        sheet.page_margins.right = margin_right_inches
        sheet.page_margins.top = margin_top_inches
        sheet.page_margins.bottom = margin_bottom_inches
        sheet.page_margins.header = margin_header_inches
        sheet.page_margins.footer = margin_footer_inches


    # --- Configurar la respuesta HTTP para la descarga ---
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    filename = f"Solicitud_{pedido_id}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response